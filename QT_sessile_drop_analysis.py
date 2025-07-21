from binascii import Error
from PyQt5 import QtWidgets, uic, QtGui
from PyQt5.QtCore import pyqtSignal
import glob, os
import cv2
from pandas.core import frame
import pyqtgraph as pg
import sys
from pathlib import Path
from skimage.filters import threshold_otsu
from edge_detection import linear_subpixel_detection as linedge
from edge_analysis import analysis
import pandas as pd
import numpy as np
import threading
import datetime
from time import sleep
import magic
import FrameSupply
from threading import Thread
from openpyxl import load_workbook
import pytz
import warnings

pg.setConfigOptions(imageAxisOrder='row-major')

filetypemap = {'image/tiff':FrameSupply.ImageReader,'image/jpeg':FrameSupply.ImageReader,'image/png':FrameSupply.ImageReader,'video/x-msvideo':FrameSupply.OpencvReadVideo}

# don't print frequent warnings to keep output clear
warnings.filterwarnings(action='ignore', message='Mean of empty slice')
warnings.filterwarnings(action='ignore', message='invalid value encountered in double_scalars')
warnings.filterwarnings(action='ignore', message='Polyfit may be poorly conditioned')

class com_handle(Thread):
    def run(self):
        self.gray = [1,2]
        self.thresh = 50
        while True:
            if len(self.gray) == 2:
                pass
            else:
                ret,thresh1 = cv2.threshold(self.gray,self.thresh,255,cv2.THRESH_BINARY)

                cv2.startWindowThread()
                cv2.namedWindow("preview")
                cv2.imshow('preview', thresh1)
                cv2.waitKey()

class MainWindow(QtWidgets.QMainWindow):
    updateVideo = pyqtSignal(np.ndarray)
    updateLeftEdge = pyqtSignal(np.ndarray,np.ndarray)
    updateRightEdge = pyqtSignal(np.ndarray,np.ndarray)
    updateLeftEdgeNeedle = pyqtSignal(np.ndarray,np.ndarray)
    updateRightEdgeNeedle = pyqtSignal(np.ndarray,np.ndarray)
    updatePlotLeft = pyqtSignal(np.ndarray,np.ndarray)
    updatePlotRight = pyqtSignal(np.ndarray,np.ndarray)
    updatePlotDeflection = pyqtSignal(np.ndarray,np.ndarray)
    updateBaseLine = pyqtSignal(np.ndarray,np.ndarray)
    updateLeftEdgeFit = pyqtSignal(np.ndarray,np.ndarray)
    updateRightEdgeFit = pyqtSignal(np.ndarray,np.ndarray)
    updateFrameCount=pyqtSignal(int,int)

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        uic.loadUi('Mainwindow.ui', self)
        self.setWindowIcon(QtGui.QIcon('icon.ico'))
        
        self.RootVidPlot=self.VideoWidget.getPlotItem()
        self.RootVidPlot.setAspectLocked(True)
        self.RootVidPlot.hideAxis('bottom')
        self.RootVidPlot.hideAxis('left')
        self.RootVidPlot.invertY(True)
        
        self.VideoItem = pg.ImageItem()
        self.RootVidPlot.addItem(self.VideoItem)
        self.BaseLineItem=pg.PlotCurveItem(pen=pg.mkPen(color='#8c564b', width=4))
        self.LeftEdgeItem=pg.PlotCurveItem(pen=pg.mkPen(color='#ff7f0e', width=2))
        self.RightEdgeItem=pg.PlotCurveItem(pen=pg.mkPen(color='#1f77b4', width=2))
        self.LeftEdgeNeedleItem=pg.PlotCurveItem(pen=pg.mkPen(color='#2ca02c', width=4))
        self.RightEdgeNeedleItem=pg.PlotCurveItem(pen=pg.mkPen(color='#2ca02c', width=4))
        self.LeftEdgeFit=pg.PlotCurveItem(pen=pg.mkPen(color='#ff7f0e', width=4))
        self.RightEdgeFit=pg.PlotCurveItem(pen=pg.mkPen(color='#1f77b4', width=4))
        
        self.RootVidPlot.addItem(self.BaseLineItem)
        self.RootVidPlot.addItem(self.LeftEdgeItem)
        self.RootVidPlot.addItem(self.RightEdgeItem)
        self.RootVidPlot.addItem(self.LeftEdgeNeedleItem)
        self.RootVidPlot.addItem(self.RightEdgeNeedleItem)
        self.RootVidPlot.addItem(self.RightEdgeFit)
        self.RootVidPlot.addItem(self.LeftEdgeFit)
        self.updateVideo.connect(self.VideoItem.setImage)
        self.updateBaseLine.connect(self.BaseLineItem.setData)
        self.updateLeftEdge.connect(self.LeftEdgeItem.setData)
        self.updateRightEdge.connect(self.RightEdgeItem.setData)
        self.updateLeftEdgeNeedle.connect(self.LeftEdgeNeedleItem.setData)
        self.updateRightEdgeNeedle.connect(self.RightEdgeNeedleItem.setData)
        self.updateLeftEdgeFit.connect(self.LeftEdgeFit.setData)
        self.updateRightEdgeFit.connect(self.RightEdgeFit.setData)
        
        self.ThetaLeftPlot=pg.ScatterPlotItem(pen='#ff7f0e',brush='#ff7f0e',symbol='o')
        self.ThetaRightPlot=pg.ScatterPlotItem(pen='#1f77b4',brush='#1f77b4',symbol='o')
        self.DeflectionPlot=pg.ScatterPlotItem(pen='#2ca02c',brush='#2ca02c',symbol='o')
        self.PlotItem=self.PlotWidget.getPlotItem()
        self.PlotItem2=self.PlotWidget2.getPlotItem()
        self.updatePlotLeft.connect(self.ThetaLeftPlot.setData)
        self.updatePlotRight.connect(self.ThetaRightPlot.setData)
        self.updatePlotDeflection.connect(self.DeflectionPlot.setData)
        
        self.BaseLine=pg.LineSegmentROI([(15,90),(100,90)],pen='#d62728')
        self.CropRoi=pg.RectROI([10,10],[110,110],scaleSnap=True)
        self.CropRoi.addScaleHandle([0,0],[1,1])
        self.VideoWidget.addItem(self.CropRoi)
        self.VideoWidget.addItem(self.BaseLine)
        self.BaseLine.addScaleHandle([15,90], [60,90])
        self.BaseLine.addRotateHandle([100,90], [15,90])
        self.BaseLine.removeHandle(0)
        self.BaseLine.removeHandle(0)
        
        self.actionNewMeasurement.triggered.connect(self.NewMeasurement)
        self.actionOpen.triggered.connect(self.openCall)
        self.actionLoadDataCard.triggered.connect(self.ReadDatacard)
        self.actionSave.triggered.connect(self.SaveResult)
        self.StartStopButton.clicked.connect(self.StartStop)
        self.CameraToggleButton.clicked.connect(self.CameraToggle)

        self.pB_FrameNumber.clicked.connect(self.SetFrameNumber)
        self.pB_FrameRotation.clicked.connect(self.SetFrameRotation)
        self.Btn_Measure_Needle.clicked.connect(self.SetNeedleDiameter)
        self.pB_NeedleDiameterMM.clicked.connect(self.SetNeedleDiameterMM)
        self.pB_PixelCalibration.clicked.connect(self.SetPixelCalibration)
        self.pB_SpringConstant.clicked.connect(self.SetSpringConstant)
        self.pB_NeedleLength.clicked.connect(self.SetNeedleLength)

        self.ModeInputCombobox.insertItems(0, ("Sessile Drop", "DAFI", "DAFI - force", "DAFI - drop width"))
        self.ModeInputCombobox.currentIndexChanged.connect(self.SetExperimentMode)
        self.StepsizeInputSpinbox.valueChanged.connect(self.SetStepsize)
        self.NStepsInputSpinbox.valueChanged.connect(self.SetNSteps)
        self.NRunsInputSpinbox.valueChanged.connect(self.SetNRuns)
        self.StepResolutionInputDoubleSpinbox.valueChanged.connect(self.SetStepResolution)
        self.kInputSpinbox.setValue(self.kInputSlider.value())
        self.kInputSlider.valueChanged.connect(lambda: self.kInputSpinbox.setValue(self.kInputSlider.value()))
        self.kInputSpinbox.valueChanged.connect(lambda: self.kInputSlider.setValue(self.kInputSpinbox.value()))
        self.updateFrameCount.connect(lambda f,n: self.FrameCounterText.setText('Frame: '+str(f)+'/'+str(n)))
        
        self.FrameSource=FrameSupply.FrameSupply()
        self.columns=['CA_L', 'CA_R', 'contactpointleft','contactpointright','leftcontact_y','rightcontact_y','volume','time','abs_time', 'fixed_base_line', 'drop_reflection', 'reflecting_edge', 'res_right', 'res_left']
        self.MeasurementResult=pd.DataFrame(columns=self.columns)

        # default analysis parameters
        self.mode = 0
        self.stepsize = 200
        self.nsteps = None
        self.nruns = 2
        self.step_resolution = 1

        self.ModeInputCombobox.setCurrentIndex(self.mode)
        self.SurfaceReflectionInputCheckbox.setCheckState(2)
        self.FixedBaselineInputCheckbox.setCheckState(0)
        self.DropReflectionInputCheckbox.setCheckState(0)
        self.StepsizeInputSpinbox.setValue(self.stepsize)
        self.NStepsInputSpinbox.setValue(0)
        self.NRunsInputSpinbox.setValue(self.nruns)
        self.StepResolutionInputDoubleSpinbox.setValue(self.step_resolution)

        # counting variables for analysis
        self.run_no = 1
        self.step_no = 0
        self.first_frame = 0
        self.PO_left = 2
        self.PO_right = 2
    

    def closeEvent(self, event):
        if self.FrameSource.is_running:
            self.FrameSource.stop()

    
    def errorpopup(self, text):
        errorpopup = QtWidgets.QMessageBox()
        errorpopup.setText(text)
        errorpopup.setStandardButtons(QtWidgets.QMessageBox.Ok)
        errorpopup.exec_()

    
    def SetExperimentMode(self):
        self.mode = int(self.ModeInputCombobox.currentIndex())
        if self.mode == 0:
            self.columns=['CA_L', 'CA_R', 'contactpointleft','contactpointright','leftcontact_y','rightcontact_y','volume','time','abs_time', 'fixed_base_line', 'drop_reflection', 'reflecting_edge', 'res_right', 'res_left']
        elif self.mode == 1:
            self.columns=['CA_L', 'CA_R', 'contactpointleft','contactpointright','leftcontact_y','rightcontact_y','deflection','time','abs_time', 'fixed_base_line', 'drop_reflection', 'reflecting_edge', 'res_right', 'res_left']
        elif self.mode == 2:
            self.columns=['deflection','time','abs_time']
        else:
            self.columns=['contactpointleft','contactpointright','leftcontact_y','rightcontact_y','width','time','abs_time', 'fixed_base_line', 'drop_reflection', 'reflecting_edge', 'res_right', 'res_left']
        self.MeasurementResult=pd.DataFrame(columns=self.columns)

      
    def SetNeedleDiameter(self):
        """Measure needle width for scale conversion."""
        if not hasattr(self, 'NeedleBox'): # create needle box that can be positioned by user
            self.NeedleBox=pg.RectROI([20,20],[80,80],scaleSnap=True)
            self.NeedleBox.addScaleHandle([0,0],[1,1])
            self.VideoWidget.addItem(self.NeedleBox)
        
        elif not hasattr(self, 'needle_calibrated'): # measure width (and rotation) of needle
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe(step=0) # get current frame
            # get crop and save coordinate transformation for needle box
            cropcoords = self.NeedleBox.getArraySlice(org_frame, self.VideoItem, returnSlice=False)
            verticalCropOffset = 0.5+cropcoords[0][0][0]
            horizontalCropOffset = 0.5+cropcoords[0][1][0]
            cropped = org_frame[cropcoords[0][0][0]:cropcoords[0][0][1],cropcoords[0][1][0]:cropcoords[0][1][1]]
            if len(org_frame.shape)==3:
                gray = cv2.cvtColor(cropped, cv2.COLOR_RGB2GRAY)
            else:
                gray = np.asarray(cropped)
            thresh = threshold_otsu(gray,np.iinfo(type(gray.flat[0])).max)
            CroppedEdgeLeft, CroppedEdgeRight = linedge(gray, thresh, mode=1)
            
            # linear fits for edges of needle
            left_b, left_a = np.polyfit(np.arange(len(CroppedEdgeLeft)), CroppedEdgeLeft, 1)
            right_b, right_a = np.polyfit(np.arange(len(CroppedEdgeRight)), CroppedEdgeRight, 1)
            delta_h = np.abs((left_b-right_b) * len(CroppedEdgeLeft)/2 + left_a-right_a) # horizontal distance between fitted linear functions (at center of crop)
            self.WidthNeedle = delta_h * np.cos(np.arctan(np.abs(left_b+right_b)/2))
            print("Width of crop: {:.2f} pixels".format(cropcoords[0][1][1]-cropcoords[0][1][0]))
            print("Width of needle: {:.2f} pixels".format(self.WidthNeedle))
            
            if self.mode == 0: # normal analysis mode -> get drop center
                self.drop_center = horizontalCropOffset + CroppedEdgeLeft[0] + self.WidthNeedle*0.5
            elif self.mode in [1,2,3]: # dafi analysis mode -> get rotation of needle
                self.NeedleRotation = np.degrees(-(left_b+right_b)/2) # rotation to compensate non-horizontal alignment of setup
                print("Rotation of needle: {:.2f} degrees".format(self.NeedleRotation))
                org_frame = self.FrameSource.rotate_image(org_frame, self.NeedleRotation) # apply rotation
                self.updateVideo.emit(org_frame) # show rotated frame
                print("Place crop at height where deflection of needle should be measured and press 'Measure Needle' again.")
                self.needle_calibrated = True

        else: # measure diameter of glass plate and initial deflection of needle (only in dafi mode)
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe(step=0)
            org_frame = self.FrameSource.rotate_image(org_frame, self.NeedleRotation)
            cropcoords = self.NeedleBox.getArraySlice(org_frame, self.VideoItem, returnSlice=False)
            verticalCropOffset = 0.5+cropcoords[0][0][0]
            horizontalCropOffset = 0.5+cropcoords[0][1][0]
            self.needle_offset = cropcoords[0][0][1] - cropcoords[0][0][0] + 5 # distance between endpoint of needle and point of deflection measurement
            self.drop_height = int(verticalCropOffset) - 1 # height of deflection measurement (upper edge of crop)

            # get width of glass plate at end of needle
            cropped = org_frame[cropcoords[0][0][1]:cropcoords[0][0][1]+100,cropcoords[0][1][0]:cropcoords[0][1][1]] # use image section below crop for measurement of plate diameter
            if len(org_frame.shape)==3:
                gray = cv2.cvtColor(cropped, cv2.COLOR_RGB2GRAY)
            else:
                gray = np.asarray(cropped)
            thresh = threshold_otsu(gray,np.iinfo(type(gray.flat[0])).max)
            CroppedEdgeLeft, CroppedEdgeRight = linedge(gray, thresh, mode=1)
            EdgeLeft = CroppedEdgeLeft + horizontalCropOffset
            EdgeRight = CroppedEdgeRight + horizontalCropOffset
            inds_needle_left = EdgeLeft > (horizontalCropOffset+1) # region where needle edge is detected
            inds_needle_right = EdgeRight < (cropcoords[0][1][1]-1)
            inds_plate_left = EdgeLeft[inds_needle_left] < (np.min(EdgeLeft[inds_needle_left]) + 20) # region with circular plate (left edge smallest, right edge largest)
            inds_plate_right = EdgeRight[inds_needle_right] > (np.max(EdgeRight[inds_needle_right]) - 20)
            dist_horizontal = np.mean(EdgeRight[inds_needle_right][inds_plate_right]) - np.mean(EdgeLeft[inds_needle_left][inds_plate_left])
            dist_vertical = np.argmax(inds_needle_right) + np.argmax(inds_plate_right) + 0.5*len(EdgeRight[inds_needle_right][inds_plate_right]) - np.argmax(inds_needle_left) - np.argmax(inds_plate_left) - 0.5*len(EdgeRight[inds_needle_left][inds_plate_left])
            self.WidthNeedle = np.sqrt(dist_horizontal**2 + dist_vertical**2) # width of circuar plate
            print("Width of crop: {:.2f} pixels".format(cropcoords[0][1][1]-cropcoords[0][1][0]))
            print("Width of needle: {:.2f} pixels".format(self.WidthNeedle))

            # get initial position of needle
            center = int(np.mean([cropcoords[0][1][0],cropcoords[0][1][1]]))
            horizontalCropOffset_needle = 0.5 + center - 50
            pos_left = np.zeros(10)
            pos_right = np.zeros(10)
            for i in range(10): # take mean of needle position for 10 frames (1 s) to compensate small fluctuations
                org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe(step=3)
                org_frame = self.FrameSource.rotate_image(org_frame, self.NeedleRotation)
                cropped_needle = org_frame[self.drop_height-10:self.drop_height,center-50:center+50] # use only small stripe (10 pixel) at drop height
                if len(org_frame.shape)==3:
                    gray = cv2.cvtColor(cropped_needle, cv2.COLOR_RGB2GRAY)
                else:
                    gray = np.asarray(cropped_needle)
                thresh = threshold_otsu(gray,np.iinfo(type(gray.flat[0])).max)
                CroppedEdgeLeftNeedle, CroppedEdgeRightNeedle = linedge(gray, thresh, mode=1)
                EdgeLeftNeedle = CroppedEdgeLeftNeedle + horizontalCropOffset_needle
                EdgeRightNeedle = CroppedEdgeRightNeedle + horizontalCropOffset_needle
                pos_left[i] = np.mean(EdgeLeftNeedle)
                pos_right[i] = np.mean(EdgeRightNeedle)
            self.defl_left = np.mean(pos_left) # save initial position of needle edge to calculate deflection for later frames
            self.defl_right = np.mean(pos_right)
            print("Initial position of needle: left {:.2f}, right {:.2f}".format(self.defl_left, self.defl_right))
            
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe(step=-30) # return to start frame without rotation
            self.updateVideo.emit(org_frame)
            del self.needle_calibrated


    # functions to read input from analysis window
    def SetFrameNumber(self):
        try:
            self.FrameSource.Set_framenumber = int(self.TB_FrameNumber.text())
            self.first_frame = int(self.TB_FrameNumber.text())
            if self.step_no > 1: # shift step number to next relevant frame
                self.step_no = np.where((self.steps-self.first_frame)>=0)[0]
            # show selected frame
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe()
            #ret,thresh1 = cv2.threshold(org_frame,139,255,cv2.THRESH_BINARY)
            self.updateVideo.emit(org_frame)
        except:
            self.FrameSource.Set_framenumber = int(self.TB_FrameNumber.text())
            self.first_frame = int(self.TB_FrameNumber.text())
            if self.step_no > 1: # shift step number to next relevant frame
                self.step_no = np.where((self.steps-self.first_frame)>=0)[0]
            # show selected frame
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe()
            #ret,thresh1 = cv2.threshold(org_frame,139,255,cv2.THRESH_BINARY)
            self.updateVideo.emit(org_frame)

    def SetFrameRotation(self):
        try:
            if hasattr(self.FrameSource, 'FrameRotation'):
                OldFrameRotation = self.FrameSource.FrameRotation
            self.FrameSource.FrameRotation = float(self.TB_FrameRotation.text())
            if hasattr(self, 'NeedleRotation'):
                try:
                    self.NeedleRotation = self.NeedleRotation - self.FrameSource.FrameRotation + OldFrameRotation
                except:
                    self.NeedleRotation = self.NeedleRotation - self.FrameSource.FrameRotation
            # show rotated frame
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe(step=0)
            self.updateVideo.emit(org_frame)
        except:
            print('Enter float')

    def SetStepsize(self):
        self.stepsize = int(self.StepsizeInputSpinbox.value())

    def SetNSteps(self):
        self.nsteps = int(self.NStepsInputSpinbox.value())

    def SetStepResolution(self):
        self.step_resolution = float(self.StepResolutionInputDoubleSpinbox.value())

    def SetNRuns(self):
        self.nruns = int(self.NRunsInputSpinbox.value())

    def SetNeedleDiameterMM(self):
        try:
            self.needle_diameter_mm = float(self.TB_NeedleDiameterMM.text())
        except:
            print('Enter float')

    def SetPixelCalibration(self):
        try:
            self.pix_calibration = float(self.TB_PixelCalibration.text())
        except:
            print('Enter float')
    
    def SetSpringConstant(self):
        try:
            self.spring_constant = float(self.TB_SpringConstant.text())
        except:
            print('Enter float')

    def SetNeedleLength(self):
        try:
            self.needle_length_mm = float(self.TB_NeedleLength.text())
        except:
            print('Enter float')


    def next_file(self):
        """Loads next video file of video sequence"""
        if self.FrameSource.is_running:
            self.FrameSource.stop()
        cwd = os.getcwd()
        if self.Labbook:
            for i, drop in enumerate(self.drop_path):
                if i == self.current_file[0]:
                    stuff = len(self.drop_path[drop])
                    if len(self.drop_path[drop]) <= self.current_file[1]:
                        if (self.mode == 1) or (self.mode == 2): # DAFI mode
                            result_path = str(self.main_path + '.xlsx')
                        elif self.mode == 3: # DAFI - width mode
                            result_path = str(self.main_path_2 + '.xlsx')
                        else:
                            result_path = str(self.main_path.replace('01_Videos', '02_Analysis_Results'))
                            if not os.path.exists(result_path):
                                os.makedirs(result_path)
                            result_path = result_path  + '/' + str(drop) + '.xlsx'
                        self.SaveResult(result_path)
                        if hasattr(self, 'start_frame_drop'):
                            del self.start_frame_drop
                        if self.mode > 0:
                            self.ModeInputCombobox.setCurrentIndex(int(3)) # set to DAFI - width mode for next drop (video from 2nd camera)
                            self.StartStopButton.setChecked(False) # stop analysis for beginning of next drop
                            self.StartStopButton.setText('Start Measurement')
                        self.current_file = [i+1, -1]
                        self.add_cap_time = 0
                        continue
                    for j, vid_path in enumerate(self.drop_path[drop]):
                        if self.current_file[1] == -1: # continue with next drop
                            VideoFile = vid_path
                            self.current_file[1] = 1
                            self.SetExperimentMode()
                            break
                        if j == self.current_file[1]: # continue with next video of same drop
                            VideoFile = vid_path
                            self.current_file[1] = j+1
                            break
        else:
            VideoFile, _ = QtWidgets.QFileDialog.getOpenFileName(self,'Open file')

        try:
            record_time = os.path.getmtime(VideoFile)
            self.file_time = datetime.datetime.fromtimestamp(record_time)
            self.file_time = self.file_time - datetime.timedelta(hours=2)
            mimetype=magic.from_file(VideoFile,mime=True)
            if VideoFile[-3:] == 'MOV':
                mimetype = 'video/x-msvideo'
            if VideoFile[-3:] == 'avi':
                mimetype = 'video/x-msvideo'
            if VideoFile[-3:] == 'MP4':
                mimetype = 'video/x-msvideo'
            if any(mimetype in key for key in filetypemap):
                if hasattr(self.FrameSource, 'FrameRotation'):
                    rot = self.FrameSource.FrameRotation
                self.FrameSource=filetypemap[mimetype](VideoFile)
                self.FrameSource.start()
                try:
                    self.FrameSource.FrameRotation = rot
                except:
                    pass
                self.FrameSource.width, self.FrameSource.height = self.FrameSource.getframesize()
                firstframe,_=self.FrameSource.getfirstframe()
                self.VideoItem.setImage(firstframe,autoRange=True)
                if self.Labbook:
                    if self.current_file[1] == 1: # new drop
                        if self.FrameSource.VideoFile[-15:-5] == "compressed":
                            params = self.GetCompressionParameters()
                            self.stepsize = 1 # analyse every frame of compressed video
                            self.step_no = -1 # start with first frame of video for correct mapping of frametimes
                            org_frame,framecaptime,milliseconds_in_vid = self.FrameSource.getnextframe(number=-1)
            else:
                self.errorpopup('Unkown filetype')
        except:
            self.last_measurement = True

    
    def reload_file(self):
        """Loads same video file again"""
        if self.FrameSource.is_running:
            self.FrameSource.stop()
        self.FrameSource.start()
        firstframe,_=self.FrameSource.getfirstframe()
        self.VideoItem.setImage(firstframe,autoRange=True)


    def path_handle(self, path, Start_drop): 
        self.is_compressed = False
        file_endings = ["*.MOV", "*.MP4"]
        cwd = os.getcwd()
        self.main_path = path
        os.chdir(path)
        drop_path = {}
        drops = []
        drop_numbers = []
        for drop in glob.glob("*/"):
            if drop[0:5] == 'Drop_':
                drops.append(drop)
                drop_numbers.append(int(drop[5:-1]))
        drop_names = {'Drops': drops, 'Numbers':drop_numbers}
        df = pd.DataFrame(drop_names)
        df = df.sort_values(by ='Numbers')
        for index, item in df.iterrows():
            drop = item['Drops']
            os.chdir(drop)
            drop_path[drop[0:-1]] = []
            for ending in file_endings:
                for video in glob.glob(ending):
                    if video[-15:-5] == "compressed":
                        self.is_compressed = True
                    drop_path[drop[0:-1]].append(os.path.join(os.getcwd(), video))
            os.chdir('..')
        self.drop_path = drop_path
        os.chdir(cwd)
        self.openCall(Start_drop)

    def path_handle_dafi(self, path, path_2):   
        self.is_compressed = False     
        cwd = os.getcwd()
        name, ending = os.path.splitext(path)
        if ("_1"+ending) in path:
            self.main_path = path.replace("_1"+ending, "")
        else:
            self.main_path = path.replace(ending, "")
        os.chdir(os.path.dirname(path))
        drop_path = {}
        drop_path["Drop_1"] = [path]
        drop_end = False
        vid_no = 1
        while not drop_end:
            vid_no += 1
            next_vid = self.main_path + "_" + str(vid_no) + ending
            if os.path.isfile(next_vid):
                drop_path["Drop_1"].append(next_vid)
            else:
                drop_end = True

        # same for videos from 2nd camera
        if path_2 != 'None':
            name, ending = os.path.splitext(path_2)
            if ("_1"+ending) in path_2:
                self.main_path_2 = path_2.replace("_1"+ending, "")
            else:
                self.main_path_2 = path_2.replace(ending, "")
            os.chdir(os.path.dirname(path_2))
            drop_path["Drop_2"] = [path_2]
            drop_end = False
            vid_no = 1
            while not drop_end:
                vid_no += 1
                next_vid = self.main_path_2 + "_" + str(vid_no) + ending
                if os.path.isfile(next_vid):
                    drop_path["Drop_2"].append(next_vid)
                else:
                    drop_end = True

        self.ModeInputCombobox.setCurrentIndex(int(1)) # set analysis mode to DAFI
        self.StepsizeInputSpinbox.setValue(int(20)) # set stepsize to 20 frames
        self.NRunsInputSpinbox.setValue(int(1)) # set number of analysis runs to 1
        Start_drop = "Drop_1"
        self.drop_path = drop_path
        print(self.drop_path)
        os.chdir(cwd)
        self.openCall(Start_drop)


    def openCall(self, Start_drop):
        self.last_measurement = False
        if self.FrameSource.is_running:
            self.FrameSource.stop()
        self.vid_num = 0
        if self.Labbook:
            if Start_drop == 'All':
                VideoFile = self.drop_path['Drop_1'][0]
                self.current_file = [0,1]
            else:
                VideoFile = self.drop_path[Start_drop][0]
                self.current_file = [int(Start_drop[5:])-1,1]
        else:
            VideoFile, _ = QtWidgets.QFileDialog.getOpenFileName(self,'Open file')
        record_time = os.path.getmtime(VideoFile)
        self.file_time = datetime.datetime.fromtimestamp(record_time)
        self.file_time = self.file_time - datetime.timedelta(hours=2)
        self.add_cap_time = 0
        mimetype = magic.from_file(VideoFile,mime=True)
        if VideoFile[-3:] == 'MOV':
            mimetype = 'video/x-msvideo'
        if VideoFile[-3:] == 'avi':
            mimetype = 'video/x-msvideo'
        if VideoFile[-3:] == 'MP4':
            mimetype = 'video/x-msvideo'
        self.MeasurementResult = pd.DataFrame(columns=self.columns)
        self.PlotItem.clear()
        self.PlotItem2.clear()
        if any(mimetype in key for key in filetypemap):
            self.FrameSource = filetypemap[mimetype](VideoFile)
            self.FrameSource.start()
            self.FrameSource.width, self.FrameSource.height = self.FrameSource.getframesize()
            self.CropRoi.setPos([self.FrameSource.width*.1,self.FrameSource.height*.1])
            self.CropRoi.setSize([self.FrameSource.width*.8,self.FrameSource.height*.8])
            self.BaseLine.setPos([self.FrameSource.width*.2,self.FrameSource.height*.7])
            firstframe,_=self.FrameSource.getfirstframe()
            self.VideoItem.setImage(firstframe,autoRange=True)
            print("Number of frames in videofile:", self.FrameSource.nframes)
        else:
            self.errorpopup('Unkown filetype')

    
    def NewMeasurement(self):
        self.StartFirstRun()
        self.MeasurementResult = pd.DataFrame(columns=self.columns)
        self.PlotItem.clear()
        self.PlotItem2.clear()
        
        
    def StartStop(self):
        if self.StartStopButton.isChecked():
            self.StartStopButton.setText('Stop Measurement')
            if self.mode == 3:
                self.PlotItem.setLabel('left', text='Contactpoint [pix]')
                self.PlotItem2.setLabel('left', text='Drop width [pix]')
            else:
                self.PlotItem.setLabel('left',text='Contact angle [°]')
                self.PlotItem2.setLabel('left',text='Deflection [pix]')
            if self.FrameSource.gotcapturetime:
                self.PlotItem.setLabel('bottom',text='Time [s]')
                self.PlotItem2.setLabel('bottom',text='Time [s]')
            else:
                self.PlotItem.setLabel('bottom',text='Frame number')
                self.PlotItem2.setLabel('bottom',text='Frame number')
            AnalysisThread = threading.Thread(target=self.RunAnalysis)
            AnalysisThread.start()
        elif not self.StartStopButton.isChecked():
            self.StartStopButton.setText('Start Measurement')
    

    def CameraToggle(self):
        if self.CameraToggleButton.isChecked():
            self.FrameSource=FrameSupply.OpencvCamera()
            self.FrameSource.start()
            FrameWidth,FrameHeight=self.FrameSource.getframesize()
            self.CropRoi.setPos([FrameWidth*.1,FrameHeight*.1])
            self.CropRoi.setSize([FrameWidth*.8,FrameHeight*.8])
            self.BaseLine.setPos([FrameWidth*.2,FrameHeight*.7])
            CameraThread = threading.Thread(target=self.CameraCapture)
            CameraThread.start()
            self.MeasurementResult=pd.DataFrame(columns=self.columns)
            self.PlotItem.clear()
            self.PlotItem2.clear()
    

    def CameraCapture(self):
        while self.CameraToggleButton.isChecked():
            if self.StartStopButton.isChecked():
                sleep(0.5)
            else:
                org_frame, _, _ = self.FrameSource.getnextframe()
                if not np.all(org_frame==-1):
                    self.updateVideo.emit(org_frame)
                else:
                    sleep(0.001)
        self.FrameSource.stop()

    def GetCompressionParameters(self):
        """Reads actual framenumbers and datetimes of compressed video from text file"""
        comp_level = int(self.FrameSource.VideoFile[-5])
        print("Analysing compressed video of compression level {}".format(comp_level))
        path = self.analysis_path[:-11]
        #path = self.FrameSource.VideoFile[:-23]+".xlsx"
        s = self.FrameSource.VideoFile
        file_name = s[s.find('Drop'):s.find('Drop')+s[s.find('Drop'):].find('\\')]+".xlsx"
        path = os.path.join(path, file_name)
        params = pd.read_excel(path, header=0)
        relevant_params = params[params['weight'] >= comp_level]
        relevant_params = relevant_params.reset_index(drop=True)
        print('Caution you are about to analyse a compsresed video, this might lead to data loss!')
        return relevant_params


    def GetWeights(self):
        """Calculates importance of every frame from changes in analysed parameters for later compression and add column to self.MeasurementResult"""
        framenumber = self.MeasurementResult['framenumber'].to_numpy()
        # parameters that are available in selected analysis mode
        if self.mode == 0:
            var = ['CA_L', 'CA_R', 'contactpointleft', 'contactpointright']
        elif self.mode == 1:
            var = ['CA_L', 'CA_R', 'contactpointleft', 'contactpointright', 'deflection']
        elif self.mode == 2:
            var = ['deflection']
        else:
            var = ['contactpointleft', 'contactpointright']
        values = self.MeasurementResult[var].to_numpy()
        tot_change = np.amax(values, axis=0) - np.min(values, axis=0) # total difference between maximal and minimal measurement values
        val_weight = tot_change[0] / tot_change # normalisation to change of left contact angle
        changes = np.abs(values[1:]-values[:-1]) # differences between contact angles / contact lines for analysed frames
        dist = np.linalg.norm(val_weight*changes, axis=1, ord=np.inf) # norm of difference vector (weighted with total change for each dimension)
        dist = np.append(dist, dist[-1]) # include last frame

        weight = np.zeros(len(framenumber))
        cum_change2 = 0
        cum_change3 = 0
        last2 = 0
        last3 = 0
        for i in range(len(framenumber)):
            cum_change2 += dist[i]
            cum_change3 += dist[i]
            if cum_change3 > 5 or (framenumber[i]-last3) > 500: # large relative change or long time since last frame of high importance
                last3 = framenumber[i]
                last2 = last3
                cum_change3 = 0
                cum_change2 = 0
                weight[i] = 3
            elif cum_change2 > 2 or (framenumber[i]-last2) > 250:
                last2 = framenumber[i]
                cum_change2 = 0
                weight[i] = 2
            else:
                weight[i] = 1
        self.MeasurementResult['weight'] = weight


    def GetFramenumbers(self):
        """Calculates framenumbers for following analysis run from changes in measured values from previous run"""
        self.steps = []
        framenumbers = self.MeasurementResult['time'].to_numpy()
        # parameters that are available in selected analysis mode
        if self.mode == 0:
            var = ['CA_L', 'CA_R', 'contactpointleft', 'contactpointright']
        elif self.mode == 1:
            var = ['CA_L', 'CA_R', 'contactpointleft', 'contactpointright', 'deflection']
        elif self.mode == 2:
            var = ['deflection']
        else:
            var = ['contactpointleft', 'contactpointright']

        values = self.MeasurementResult[var].to_numpy()

        if self.mode != 2:#filter out bad ca measurements
            residuals = self.MeasurementResult[['res_right', 'res_left']].to_numpy()
            true_f = (residuals<0.0001) + (residuals>1000)
            true_f = np.invert((true_f[:,0]) & (true_f[:,1]))#if both are false (bad fit on left and right cl)
            framenumbers = framenumbers[true_f]
            values = values[true_f]


        # use only current video of drop (exclude measurement results fom previous videos)
        values = values[framenumbers>=self.add_cap_time]
        framenumbers = framenumbers[framenumbers>=self.add_cap_time]

        if  values.shape[0] < 2: # in case of very short end video
            tot_change = np.asarray([1,1,1,1])
        else:
           tot_change = np.amax(values, axis=0) - np.min(values, axis=0) # total difference between maximal and minimal measurement values
        weight = tot_change[0] / tot_change # normalisation to change of left contact angle
        changes = np.abs(values[1:]-values[:-1]) # differences between contact angles / contact lines for analysed frames
        dist = np.linalg.norm(changes*weight, axis=1, ord=np.inf) # norm of difference vector (weighted with total change for each dimension)
        n = np.ceil(dist/self.step_resolution) # number of frames to analyse in each intervall (ceil n for at least one frame)
        minimal_stepsize = 1 #minimal intervall via input settigs
        n = np.min((n, (framenumbers[1:]-framenumbers[:-1])/minimal_stepsize), axis=0) # maximum number = total number of frames in intervall

        for i in range(len(framenumbers)-1):
            # new famenumbers evenly spaced in intervall between old framenumbers, remove start number (already analysed)
            new_framenumbers = np.linspace(framenumbers[i], framenumbers[i+1], int(n[i]), endpoint=False)[1:]
            self.steps = np.concatenate((self.steps, np.round(new_framenumbers))) # append framenumbers rounded to integer values
        
        if len(framenumbers)>3:#case of very short video will be ignored
            # include frames before first frame of previous analysis
            new_framenumbers = np.linspace(2*framenumbers[0]-framenumbers[1], framenumbers[0], int(n[0]), endpoint=False)[1:]
            self.steps = np.concatenate((np.round(new_framenumbers), self.steps))
            # include frames after last frame of previous analysis
            new_framenumbers = np.linspace(framenumbers[-1], 2*framenumbers[-1]-framenumbers[-2], int(n[-1]), endpoint=False)[1:]
            self.steps = np.concatenate((self.steps, np.round(new_framenumbers)))
            # # continue with same stepsize for frames with bad fit quality at end of video
            # step = int(self.steps[-1] - self.steps[-2])
            # last_framenumber = 2*self.MeasurementResult['time'].to_numpy()[-1]-self.MeasurementResult['time'].to_numpy()[-2]
            # new_framenumbers = np.arange(2*framenumbers[-1]-framenumbers[-2], last_framenumber, step)[1:]
            # self.steps = np.concatenate((self.steps, np.round(new_framenumbers)))

        self.steps = np.array(self.steps) - self.add_cap_time # start counting framenumbers with 0 for each video
        self.steps = np.concatenate((self.steps, [int(self.FrameSource.nframes)])) # add last framenumber to skip end of video when drop is gone

        print("Frame numbers to be analysed in the next run: ", self.steps)
        print("Number of frames to be analysed in the next run: ", len(self.steps))


    def GetBaseline(self, org_frame, horizontalCropOffset, verticalCropOffset):
        """Reads baseline positions from window and extrapolates to the edge of the crop"""
        _,basearray = self.BaseLine.getArrayRegion(org_frame, self.VideoItem, returnSlice=False, returnMappedCoords=True)
        baseinput = [[basearray[0,0], basearray[1,0]-verticalCropOffset], [basearray[0,-1], basearray[1,-1]-verticalCropOffset]]
        del basearray
        rightbasepoint = np.argmax([baseinput[0][0], baseinput[1][0]])
        baseslope = float(baseinput[rightbasepoint][1]-baseinput[1-rightbasepoint][1])/(baseinput[rightbasepoint][0]-baseinput[1-rightbasepoint][0])
        base = np.array([[0, baseinput[0][1]-baseslope*baseinput[0][0]], [org_frame.shape[1], baseslope*org_frame.shape[1]+baseinput[0][1]-baseslope*baseinput[0][0]]])
        return base
    
    
    def UpdateWindow(self, verticalCropOffset, EdgeLeft=[], EdgeRight=[], plotinfo=[], base=[], EdgeLeftNeedle=[], EdgeRightNeedle=[]):
        """Shows current frame in window and adds current measurement values to plot"""
        if self.FrameSource.gotcapturetime:
            plottime = self.MeasurementResult['time']-self.MeasurementResult.iloc[0]['time']
            plottime = plottime.to_numpy().astype('float')*10**-9 # convert from nanoseconds to seconds
        else:
            plottime = self.MeasurementResult['time'].to_numpy()
        self.updateFrameCount.emit(int(self.FrameSource.framenumber), int(self.FrameSource.nframes))

        if self.mode in [0,1,3]:
            self.updateBaseLine.emit(base[:,0], base[:,1]+verticalCropOffset)
            self.updateLeftEdge.emit(EdgeLeft, np.arange(0,len(EdgeLeft))+verticalCropOffset)
            self.updateRightEdge.emit(EdgeRight, np.arange(0,len(EdgeRight))+verticalCropOffset)
            self.updateLeftEdgeFit.emit(plotinfo[0,:], verticalCropOffset+plotinfo[1,:])
            self.updateRightEdgeFit.emit(plotinfo[2,:], verticalCropOffset+plotinfo[3,:])

        if self.mode in [0,1]:
            plotleft = self.MeasurementResult['CA_L'].to_numpy()
            plotright = self.MeasurementResult['CA_R'].to_numpy()
            self.updatePlotLeft.emit(plottime, plotleft)
            self.updatePlotRight.emit(plottime, plotright)

        if self.mode in [1,2]:
            deflection = self.MeasurementResult['deflection'].to_numpy()
            # https://docs.opencv.org/2.4/modules/core/doc/operations_on_arrays.html?highlight=perspectivetransform#perspectivetransform
            self.updateLeftEdgeNeedle.emit(EdgeLeftNeedle, np.arange(0,len(EdgeLeftNeedle))+self.drop_height-10)
            self.updateRightEdgeNeedle.emit(EdgeRightNeedle, np.arange(0,len(EdgeRightNeedle))+self.drop_height-10)
            self.updatePlotDeflection.emit(plottime, deflection)

        if self.mode == 3:
            plotleft = self.MeasurementResult['contactpointleft'].to_numpy()
            plotright = self.MeasurementResult['contactpointright'].to_numpy()
            width = self.MeasurementResult['width'].to_numpy()
            self.updatePlotLeft.emit(plottime, plotleft)
            self.updatePlotRight.emit(plottime, plotright)
            self.updatePlotDeflection.emit(plottime, width)


    def StartFirstRun(self):
        """Prepare first analysis run of video file"""
        self.run_no = 1
        if not hasattr(self, 'start_frame_drop'):
            self.start_frame_drop = self.first_frame
        self.first_frame = 0
        self.add_cap_time += self.FrameSource.nframes
        if self.mode == 0: # gap of 250 s between videos in normal mode
            self.add_cap_time += 250
        try:
            self.MeasurementResult = self.MeasurementResult.sort_values(by=['time'])
            self.MeasurementResult = self.MeasurementResult.reset_index(drop=True)
        except:
            pass
        self.steps = []
        self.step_no = 0
        print("Start first analysis run of next file.")

    
    def StartSecondRun(self):
        """Prepare second analysis run of video file"""
        self.run_no +=1
        self.MeasurementResult = self.MeasurementResult.sort_values(by=['time'])
        self.MeasurementResult = self.MeasurementResult.reset_index(drop=True)
        self.GetFramenumbers()
        self.FrameSource.framenumber = self.first_frame + self.add_cap_time
        self.step_no = 0
        print("Start analysis run {} for same file.".format(self.run_no))


    def RunAnalysis(self):
        """Calculation of contactpoints and contactangles from the videoframes."""
        self.PlotItem.addItem(self.ThetaLeftPlot)
        self.PlotItem.addItem(self.ThetaRightPlot)
        self.PlotItem2.addItem(self.DeflectionPlot)

        if self.nsteps: # adjust stepsize to constant number of analysed frames
            self.stepsize = int(self.FrameSource.nframes/self.nsteps)

        if self.FrameSource.VideoFile[-15:-5] == "compressed":
            params = self.GetCompressionParameters()
            self.stepsize = 1 # analyse every frame of compressed video
            self.step_no = -1 # start with first frame of video for correct mapping of frametimes
            org_frame,framecaptime,milliseconds_in_vid = self.FrameSource.getnextframe(number=-1)

        while self.StartStopButton.isChecked(): # analysis loop for video frames
            try: # analysis with given variable framerate (not first analysis run of video or already compressed video)
                org_frame,framecaptime,milliseconds_in_vid = self.FrameSource.getnextframe(number=self.steps[self.step_no])
            except: # analysis with constant framerate (first analysis run of video)
                org_frame,framecaptime,milliseconds_in_vid = self.FrameSource.getnextframe(step=self.stepsize)
            self.step_no += 1

            if not np.all(org_frame==-1): # frame was loaded successfully
                # get crop and save coordinate transformation
                abs_frame_time = self.file_time+datetime.timedelta(milliseconds=milliseconds_in_vid)
                self.updateVideo.emit(org_frame)
                cropcoords=self.CropRoi.getArraySlice(org_frame, self.VideoItem, returnSlice=False)
                verticalCropOffset=0.5+cropcoords[0][0][0]
                horizontalCropOffset=0.5+cropcoords[0][1][0]
                try:
                    drop_center = int(self.drop_center - horizontalCropOffset)
                except:
                    drop_center = int(np.mean([cropcoords[0][1][0], cropcoords[0][1][1]]) - horizontalCropOffset)
                frame_time = framecaptime + self.add_cap_time
                if self.FrameSource.VideoFile[-15:-5] == "compressed":
                    frame_time = params['framenumber'][self.step_no]
                    abs_frame_time = params['abs_time'][self.step_no]

                try:
                    if self.mode != 2: # get contact points and contact angles
                        cropped = org_frame[cropcoords[0][0][0]:cropcoords[0][0][1],cropcoords[0][1][0]:cropcoords[0][1][1]]
                        if len(org_frame.shape)==3:
                            gray = cv2.cvtColor(cropped, cv2.COLOR_RGB2GRAY)
                        else:
                            gray = np.asarray(cropped)
                        thresh = threshold_otsu(gray,np.iinfo(type(gray.flat[0])).max)
                        if self.DropReflectionInputCheckbox.checkState(): # light edge of drop due to reflection -> repeat threshold calculation
                            gray_edge = gray
                            for i in range(2):
                                gray_edge = gray_edge[gray_edge>thresh]
                                thresh = threshold_otsu(gray_edge,np.iinfo(type(gray_edge.flat[0])).max)
                        """ com = com_handle()
                        com.start()
                        com.gray = gray
                        com.thresh = thresh
                        del com """
                        try:
                            # time between frames small enough, deviation of measurements small enough, enough measurements performed already
                            ReliableMeasurement = (abs(old_frame-(frame_time))<500) & (len(self.MeasurementResult)>6)
                            if self.mode == 0:
                                std_derev  = self.MeasurementResult['CA_L'][-5:].std() + self.MeasurementResult['CA_R'][-5:].std()
                                std_derev_pos  = self.MeasurementResult['contactpointleft'][-5:].std() + self.MeasurementResult['contactpointright'][-5:].std()
                                ReliableMeasurement = ReliableMeasurement & (std_derev<10) & (std_derev_pos<13)
                            if ReliableMeasurement: # use drop edge from previous frame as orientation for new edge
                                CroppedEdgeLeft, CroppedEdgeRight = linedge(gray, thresh, drop_center, old_left_arr, old_right_arr, mode=self.mode)
                            else:
                                raise Error('Do not trust previous measurements')
                        except: # search for drop edge in entire frame
                            CroppedEdgeLeft, CroppedEdgeRight = linedge(gray, thresh, drop_center, mode=self.mode)
                        EdgeLeft = CroppedEdgeLeft + horizontalCropOffset
                        EdgeRight = CroppedEdgeRight + horizontalCropOffset
                        
                        if self.FixedBaselineInputCheckbox.checkState() == 2:
                            fixed_base_line = True
                            base = self.GetBaseline(org_frame, horizontalCropOffset, verticalCropOffset)
                            contactpointleft, contactpointright, thetal, thetar, dropvolume, plotinfo, self.PO_left, self.PO_right, leftcontact_y, rightcontact_y, res_right, res_left = analysis(EdgeLeft, EdgeRight, base, cropped.shape, k=self.kInputSpinbox.value(), PO_left=self.PO_left, PO_right=self.PO_right, reflection=self.SurfaceReflectionInputCheckbox.checkState(), BaselineFixed= True)
                        else:
                            fixed_base_line = False
                            try: # use baseline from previous frame as orientation for new baseline
                                if (self.run_no!=1) and (abs(old_frame-(frame_time))>500): # search baseline of closest frame in results from previous analysis
                                    close_index = self.MeasurementResult['time'].sub(frame_time).astype(float).abs().idxmin()
                                    old_res_left = self.MeasurementResult['res_left'][close_index-10:close_index]
                                    old_res_right = self.MeasurementResult['res_right'][close_index-10:close_index]
                                    if (np.sum(old_res_left < 100) + np.sum(old_res_left > 0.001) + np.sum(old_res_right < 100) + np.sum(old_res_right > 0.001)) >= 38: # use baseline from previous analysis only if residual is good
                                        old_baseslope = float(self.MeasurementResult['rightcontact_y'][close_index] - self.MeasurementResult['leftcontact_y'][close_index]) / (self.MeasurementResult['contactpointright'][close_index] - self.MeasurementResult['contactpointleft'][close_index])
                                        old_base = np.array([[0, self.MeasurementResult['leftcontact_y'][close_index]-old_baseslope*self.MeasurementResult['contactpointleft'][close_index]], [org_frame.shape[1], old_baseslope*org_frame.shape[1]+self.MeasurementResult['leftcontact_y'][close_index]-old_baseslope*self.MeasurementResult['contactpointleft'][close_index]]])
                                    else: # get baseline from window
                                        raise Error
                                contactpointleft, contactpointright, thetal, thetar, dropvolume, plotinfo, self.PO_left, self.PO_right, leftcontact_y, rightcontact_y, res_right, res_left = analysis(EdgeLeft, EdgeRight, old_base, cropped.shape, k=self.kInputSpinbox.value(), PO_left=self.PO_left, PO_right=self.PO_right, reflection=self.SurfaceReflectionInputCheckbox.checkState())
                                old_baseslope = float(rightcontact_y - leftcontact_y) / (contactpointright - contactpointleft) # fitted baseline (line between contact points)
                                try: # orientation baseline = 0.8*old baseline + 0.2*current baseline
                                    if (abs(old_leftcontact_y - leftcontact_y) < 2) and (abs(old_rightcontact_y - rightcontact_y) < 2): # if height of baseline does not jump too much
                                        old_base = 0.8*old_base + 0.2*np.array([[0, leftcontact_y-old_baseslope*contactpointleft], [org_frame.shape[1], old_baseslope*org_frame.shape[1]+leftcontact_y-old_baseslope*contactpointleft]])
                                except: # orientation baseline = current baseline
                                    old_base = np.array([[0, leftcontact_y-old_baseslope*contactpointleft], [org_frame.shape[1], old_baseslope*org_frame.shape[1]+leftcontact_y-old_baseslope*contactpointleft]])
                            except: # get baseline from window
                                base = self.GetBaseline(org_frame, horizontalCropOffset, verticalCropOffset)
                                contactpointleft, contactpointright, thetal, thetar, dropvolume, plotinfo, self.PO_left, self.PO_right, leftcontact_y, rightcontact_y, res_right, res_left = analysis(EdgeLeft, EdgeRight, base, cropped.shape, k=self.kInputSpinbox.value(), PO_left=self.PO_left, PO_right=self.PO_right, reflection=self.SurfaceReflectionInputCheckbox.checkState())
                                old_base = base # orientation baseline = baseline from window

                        # orientation parameters for next drop
                        old_leftcontact_y = leftcontact_y
                        old_rightcontact_y = rightcontact_y
                        old_left_arr = CroppedEdgeLeft
                        old_right_arr = CroppedEdgeRight

                    if self.mode in [1,2]: # get deflection of needle
                        org_frame = self.FrameSource.rotate_image(org_frame, self.NeedleRotation) # rotate frame to measure needle deflection orthogonal to needle
                        try:
                            cropped_needle = org_frame[self.drop_height-10:self.drop_height,old_needle_pos-50:old_needle_pos+50] # use only small stripe (10 pixel) at drop height
                        except:
                            old_needle_pos = int(np.mean([self.defl_left, self.defl_right]))
                            cropped_needle = org_frame[self.drop_height-10:self.drop_height,old_needle_pos-50:old_needle_pos+50]
                        horizontalCropOffset_needle = 0.5 + old_needle_pos - 50
                        if len(org_frame.shape)==3:
                            gray = cv2.cvtColor(cropped_needle, cv2.COLOR_RGB2GRAY)
                        else:
                            gray = np.asarray(cropped_needle)
                        if (np.max(gray) - np.min(gray)) < 20: # needle not in search array, use width of drop crop
                            cropped_needle = org_frame[self.drop_height-10:self.drop_height,cropcoords[0][1][0]:cropcoords[0][1][1]]
                            horizontalCropOffset_needle = horizontalCropOffset
                            if len(org_frame.shape)==3:
                                gray = cv2.cvtColor(cropped_needle, cv2.COLOR_RGB2GRAY)
                            else:
                                gray = np.asarray(cropped_needle)
                        thresh = threshold_otsu(gray,np.iinfo(type(gray.flat[0])).max)
                        CroppedEdgeLeftNeedle, CroppedEdgeRightNeedle = linedge(gray, thresh, mode=1)
                        EdgeLeftNeedle = CroppedEdgeLeftNeedle + horizontalCropOffset_needle
                        EdgeRightNeedle = CroppedEdgeRightNeedle + horizontalCropOffset_needle
                        pos_left = np.mean(EdgeLeftNeedle)
                        pos_right = np.mean(EdgeRightNeedle)
                        if (pos_right - pos_left > 2*(self.defl_right - self.defl_left)):
                            old_needle_pos = int(np.mean([self.defl_left, self.defl_right]))
                            raise Error('Wrong needle position found.')
                        deflection = -np.mean([pos_left-self.defl_left, pos_right-self.defl_right])
                        old_needle_pos = int(np.mean([pos_left, pos_right]))

                    if self.mode == 0:
                        val = [thetal, thetar, contactpointleft, contactpointright, leftcontact_y, rightcontact_y, dropvolume, frame_time, abs_frame_time, fixed_base_line, self.SurfaceReflectionInputCheckbox.checkState(), self.DropReflectionInputCheckbox.checkState(), res_right, res_left]
                        self.UpdateWindow(verticalCropOffset, EdgeLeft=EdgeLeft, EdgeRight=EdgeRight, plotinfo=plotinfo, base=old_base)
                    elif self.mode == 1:
                        val = [thetal, thetar, contactpointleft, contactpointright, leftcontact_y, rightcontact_y, deflection, frame_time, abs_frame_time, fixed_base_line, self.SurfaceReflectionInputCheckbox.checkState(), self.DropReflectionInputCheckbox.checkState(), res_right, res_left]
                        self.UpdateWindow(verticalCropOffset, EdgeLeft=EdgeLeft, EdgeRight=EdgeRight, plotinfo=plotinfo, base=old_base, EdgeLeftNeedle=EdgeLeftNeedle, EdgeRightNeedle=EdgeRightNeedle)
                    elif self.mode == 2:
                        val = [deflection, frame_time, abs_frame_time]
                        self.UpdateWindow(verticalCropOffset, EdgeLeftNeedle=EdgeLeftNeedle, EdgeRightNeedle=EdgeRightNeedle)
                    else:
                        val = [contactpointleft, contactpointright, leftcontact_y, rightcontact_y, contactpointright-contactpointleft, frame_time, abs_frame_time, fixed_base_line, self.SurfaceReflectionInputCheckbox.checkState(), self.DropReflectionInputCheckbox.checkState(), res_right, res_left]
                        self.UpdateWindow(verticalCropOffset, EdgeLeft=EdgeLeft, EdgeRight=EdgeRight, plotinfo=plotinfo, base=old_base)
                    newrow = pd.DataFrame([val], columns=self.columns)
                    self.MeasurementResult = pd.concat([self.MeasurementResult, newrow], ignore_index=True)
                    old_frame = frame_time

                except:
                    self.updateFrameCount.emit(int(self.FrameSource.framenumber), int(self.FrameSource.nframes))
                    print('failed measurement for frame number {}'.format(self.FrameSource.framenumber))
            else:
                sleep(0.001)

            if (not self.FrameSource.is_running and len(self.FrameSource.framebuffer)<1): # end of video reached
                if self.last_measurement: # analysis of all drops finished
                    break
                elif (self.FrameSource.VideoFile[-15:-5] == "compressed") or (self.run_no == self.nruns): # last analysis run of video finished
                    self.StartFirstRun()
                    self.next_file()
                else: # do another analysis run for same video
                    self.StartSecondRun()
                    if len(self.steps) == 0:
                        self.StartFirstRun()
                        self.next_file()
                    else:
                        self.reload_file()

            
    def SaveResult(self, path=None):
        """Writes results from video analysis to excel file and converts values to physical units, if conversion scale is available"""
        if len(self.MeasurementResult.index)>0:
            if path == None or path == False:
                SaveFileName, _ = QtWidgets.QFileDialog.getSaveFileName(self,'Save file', '', "Excel Files (*.xlsx);;All Files (*)")
            else:
                SaveFileName = path
            if self.FrameSource.VideoFile[-15:-5] == 'compressed':
                SaveFileName = SaveFileName[:-5] + '_compressed' + self.FrameSource.VideoFile[-5] + SaveFileName[-5:]
            self.MeasurementResult = self.MeasurementResult.sort_values(by=['time'])
            self.MeasurementResult = self.MeasurementResult.reset_index(drop=True)
            if not self.FrameSource.gotcapturetime:
                self.MeasurementResult=self.MeasurementResult.rename(columns={"time": "framenumber"})
            if hasattr(self, 'WidthNeedle') and not hasattr(self, 'pix_calibration'): # convert pixel to mm using known width of needle
                if not hasattr(self, 'needle_diameter_mm'): # use default values if no needle with is provided
                    if self.mode == 0: # diameter of needle in normal setup
                        self.needle_diameter_mm = 1.829
                    elif self.mode in [1,2,3]: # diameter of glass plate
                        self.needle_diameter_mm = 5
                self.pix_calibration = self.needle_diameter_mm/self.WidthNeedle
            if hasattr(self, 'pix_calibration'): # convert measured lengths in mm
                if self.mode in [0,1,3]:
                    self.MeasurementResult['BI_left'] = self.MeasurementResult['contactpointleft']*self.pix_calibration
                    self.MeasurementResult['BI_right'] = self.MeasurementResult['contactpointright']*self.pix_calibration
                if self.mode in [1,2]:
                    self.MeasurementResult['deflection / mm'] = self.MeasurementResult['deflection']*self.pix_calibration
                if self.mode == 3:
                    self.MeasurementResult['width / mm'] = self.MeasurementResult['width']*self.pix_calibration
                if self.mode in [1,2] and hasattr(self, 'spring_constant'): # convert deflection to force using known spring constant of needle
                    if hasattr(self, 'needle_length_mm') and hasattr(self, 'needle_offset'): # convert spring constant to actual needle length
                        spring_constant = self.spring_constant*self.needle_length_mm/(self.needle_length_mm-self.needle_offset*self.pix_calibration)
                    else:
                        spring_constant = self.spring_constant
                    self.MeasurementResult['force / mN'] = self.MeasurementResult['deflection / mm']*spring_constant
                self.GetWeights()
                self.MeasurementResult.to_excel(SaveFileName)
            self.CreateDatacard(SaveFileName)
            if hasattr(self, 'pix_calibration'):
                del self.pix_calibration
        else:
            self.errorpopup('Nothing to save')


    def CreateDatacard(self, SaveFileName):
        """Writes all relevant parameters of analysis to textfile (position of measurement box, initial position of baseline, stepsize etc.)"""
        parameters = ["analysis_time", "VideoFile", "mode", "surface reflection", "drop reflection", "fixed baseline", "edge pixels", "BaseLine.pos_x", "BaseLine.pos_y", "BaseLine.size_x", "BaseLine.angle", "CropRoi.pos_x", "CropRoi.pos_y", "CropRoi.size_x", "CropRoi.size_y", "NeedleBox.pos_x", "NeedleBox.pos_y", "NeedleBox.size_x", "NeedleBox.size_y"]
        values = [datetime.datetime.now(), self.FrameSource.VideoFile, self.mode, self.SurfaceReflectionInputCheckbox.checkState(), self.DropReflectionInputCheckbox.checkState(), self.FixedBaselineInputCheckbox.checkState(), self.kInputSpinbox.value(), self.BaseLine.pos()[0], self.BaseLine.pos()[1], self.BaseLine.size()[0], self.BaseLine.angle(), self.CropRoi.pos()[0], self.CropRoi.pos()[1], self.CropRoi.size()[0], self.CropRoi.size()[1], self.NeedleBox.pos()[0], self.NeedleBox.pos()[1], self.NeedleBox.size()[0], self.NeedleBox.size()[1]]
        datacard = pd.DataFrame({'value': values}, index=parameters)
        optional_parameters = ["stepsize", "nsteps", "step_resolution", "nruns", "start_frame_drop", "PO_left", "PO_right", "WidthNeedle", "drop_center", "drop_height", "defl_left", "defl_right", "needle_offset", "NeedleRotation", "needle_diameter_mm", "pix_calibration", "spring_constant", "needle_length_mm", "FrameRotation"]
        for param in optional_parameters:
            if hasattr(self, param):
                newrow = pd.DataFrame([[getattr(self, param)]], columns=['value'], index=[param])
                datacard = pd.concat([datacard, newrow])

        book = load_workbook(SaveFileName)
        book.create_sheet('Datacard')
        writer = pd.ExcelWriter(SaveFileName, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        datacard.to_excel(writer, sheet_name='Datacard', startcol=writer.sheets['Datacard'].max_column-1, index=True)
        writer.close()

    
    def ReadDatacard(self, path = None):
        """Sets parameters of analysis to values from datacard"""
        try:
            if path == None or path == False:
                path, _ = QtWidgets.QFileDialog.getOpenFileName(self,'Select datacard')
            try:
                data = pd.read_excel(path, sheet_name='Datacard', index_col=0)
            except: # old datacard format
                data = pd.read_table(path, sep="    ", header=None, index_col=0, names=["value"], engine="python")
        except:
            print('No valid filetype for datacard')
            return
        self.analysis_path = path

        try:
            int_parameters = ["PO_left", "PO_right", "drop_height", "needle_offset"]
            float_parameters = ["WidthNeedle", "drop_center", "defl_left", "defl_right", "NeedleRotation", "needle_diameter_mm", "pix_calibration", "spring_constant", "needle_length_mm", "FrameRotation"]
            for param in int_parameters:
                try:
                    setattr(self, param, int(data["value"][param]))
                except:
                    pass
            for param in float_parameters:
                try:
                    setattr(self, param, float(data["value"][param]))
                except:
                    pass

            self.ModeInputCombobox.setCurrentIndex(int(data["value"]["mode"]))
            self.SurfaceReflectionInputCheckbox.setCheckState(int(data["value"]["surface reflection"]))
            self.DropReflectionInputCheckbox.setCheckState(int(data["value"]["drop reflection"]))
            try:
                self.FixedBaselineInputCheckbox.setCheckState(int(data["value"]["fixed baseline"]))
            except:
                self.FixedBaselineInputCheckbox.setCheckState(0)
            self.StepsizeInputSpinbox.setValue(int(data["value"]["stepsize"]))
            try:
                self.NStepsInputSpinbox.setValue(int(data["value"]["nsteps"]))
            except:
                self.NStepsInputSpinbox.setValue(0)
                self.nsteps = None
            self.NRunsInputSpinbox.setValue(int(data["value"]["nruns"]))
            self.StepResolutionInputDoubleSpinbox.setValue(float(data["value"]["step_resolution"]))
            self.kInputSpinbox.setValue(int(data["value"]["edge pixels"]))

            try:
                self.first_frame = int(data["value"]["start_frame_drop"])
            except:
                data2 = pd.read_excel(path, sheet_name='Sheet1', index_col=0)
                self.first_frame = int(data2["framenumber"].to_numpy()[0])
            if self.is_compressed:
                self.first_frame = 1
            self.FrameSource.Set_framenumber = self.first_frame
            org_frame, framecaptime, milliseconds_in_vid = self.FrameSource.getnextframe()
            self.updateVideo.emit(org_frame)

            self.BaseLine.setPos([float(data["value"]["BaseLine.pos_x"]), float(data["value"]["BaseLine.pos_y"])])
            self.BaseLine.setSize([float(data["value"]["BaseLine.size_x"]), 1])
            self.BaseLine.setAngle(float(data["value"]["BaseLine.angle"]))
            self.CropRoi.setPos([float(data["value"]["CropRoi.pos_x"]), float(data["value"]["CropRoi.pos_y"])])
            self.CropRoi.setSize([float(data["value"]["CropRoi.size_x"]), float(data["value"]["CropRoi.size_y"])])
            if not hasattr(self, 'NeedleBox'):
                self.NeedleBox=pg.RectROI([20,20],[80,80],scaleSnap=True)
                self.NeedleBox.addScaleHandle([0,0],[1,1])
                self.VideoWidget.addItem(self.NeedleBox)
            self.NeedleBox.setPos([float(data["value"]["NeedleBox.pos_x"]), float(data["value"]["NeedleBox.pos_y"])])
            self.NeedleBox.setSize([float(data["value"]["NeedleBox.size_x"]), float(data["value"]["NeedleBox.size_y"])])
        
        except:
            self.errorpopup('Could not read datacard')
    
    
def main():
    if len(sys.argv) == 1:
        Labbook = False
    else:
        Labbook = True
        os.chdir(sys.argv[3])
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon('icon.ico'))
    main = MainWindow()
    main.show()
    main.Labbook = Labbook
    if Labbook:
        if sys.argv[2] == "DAFI":
            main.path_handle_dafi(sys.argv[1], sys.argv[5])
            if sys.argv[4] == 'None':
                pass
            else:
                main.ReadDatacard(sys.argv[4])
            pass
        else:
            main.path_handle(sys.argv[1], sys.argv[2])
            if sys.argv[4] == 'None': # no datacard to read
                pass
            elif sys.argv[2] == 'All': # read datacard of first drop
                main.ReadDatacard(sys.argv[4]+'/Drop_1.xlsx')
            else: # read datacard of corresponding drop
                main.ReadDatacard(sys.argv[4]+'/'+sys.argv[2]+'.xlsx')
    sys.exit(app.exec_())

if __name__ == '__main__':         
    main()
